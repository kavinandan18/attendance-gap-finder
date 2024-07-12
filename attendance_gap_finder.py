import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def load_data(main_file, dup_file):
    """Load data from various file formats into Pandas dataframes."""
    main_df = None
    dup_df = None

    try:
        if main_file.name.endswith('.xlsx'):
            main_df = pd.read_excel(main_file)
        elif main_file.name.endswith('.csv'):
            main_df = pd.read_csv(main_file)

        if dup_file.name.endswith('.xlsx'):
            dup_df = pd.read_excel(dup_file)
        elif dup_file.name.endswith('.csv'):
            dup_df = pd.read_csv(dup_file)

        return main_df, dup_df

    except Exception as e:
        st.error(f"Error loading files: {str(e)}")
        return None, None

def create_fill(color="FFFF0000"):
    """Create a PatternFill for highlighting."""
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

def find_best_match(row_main, dup_df):
    """Find the best matching row in the duplicate dataframe."""
    best_match_idx = -1
    min_diff_count = float('inf')

    for idx_dup, row_dup in dup_df.iterrows():
        diff_count = (row_main != row_dup).sum()
        if diff_count < min_diff_count:
            min_diff_count = diff_count
            best_match_idx = idx_dup

    return best_match_idx

def highlight_differences(main_df, dup_df, main_ws, fill):
    """Highlight rows in the main worksheet where data is missing or different in the duplicate dataframe."""
    for idx_main, row_main in main_df.iterrows():
        best_match_idx = find_best_match(row_main, dup_df)

        if best_match_idx != -1:
            row_dup = dup_df.iloc[best_match_idx]
            row_different = False
            for col in main_df.columns:
                if col in dup_df.columns:
                    if pd.isna(row_dup[col]) or row_main[col] != row_dup[col]:
                        row_different = True
                        break
            if row_different:
                for col in main_df.columns:
                    cell = main_ws.cell(row=idx_main + 2, column=main_df.columns.get_loc(col) + 1)
                    cell.fill = fill

def process_files(main_file, dup_file):
    main_df, dup_df = load_data(main_file, dup_file)

    if main_df is None or dup_df is None:
        return None, None, None

    try:
        main_wb = load_workbook(main_file)
        main_ws = main_wb.active

        red_fill = create_fill("FFFF0000")

        highlight_differences(main_df, dup_df, main_ws, red_fill)

        output_file = 'Highlighted_Empmain.xlsx'
        main_wb.save(output_file)

        return output_file, main_df, dup_df

    except Exception as e:
        st.error(f"Error processing files: {str(e)}")
        return None, None, None

def main():
    st.set_page_config(page_title="Attendance Gap Finder", layout="wide")
    st.markdown("""
    <style>
    .reportview-container {
        background: #e8f5fd;
        padding: 20px;
        border-radius: 10px;
    }
    .sidebar .sidebar-content {
        background: #f0f2f6;
        border-radius: 10px;
    }
    .footer {
        font-size: 0.8rem;
        text-align: center;
        color: #fff;
        margin-top: 20px;
        background-color: #333;
        padding: 10px 0;
    }
    .center-title {
        display: flex;
        justify-content: center;
        align-items: center;
        height: 100px;
        background-color: #4CAF50; /* Title background color */
        color: white;
    }
    .center-button {
        display: flex;
        justify-content: center;
        margin-top: 20px;
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown('<div class="center-title"><h1>Attendance Gap Finder</h1></div>', unsafe_allow_html=True)

    main_file = st.file_uploader("Upload Main File", type=["xlsx", "csv"])
    dup_file = st.file_uploader("Upload Duplicate File", type=["xlsx", "csv"])

    if st.button("Find Gap", key="find_gap"):
        if main_file and dup_file:
            output_file, main_df, dup_df = process_files(main_file, dup_file)

            if output_file and main_df is not None and dup_df is not None:
                st.success("Gap Analysis Completed!")

                st.markdown("### Main File Data:")
                st.dataframe(main_df)

                st.markdown("### Duplicate File Data:")
                st.dataframe(dup_df)

                with open(output_file, "rb") as file:
                    st.download_button(label="Download Highlighted Main File", data=file, file_name=output_file, mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    st.markdown('<div class="footer">@ 2024 Avinandan Kumar</div>', unsafe_allow_html=True)

if __name__ == "__main__":
    main()
